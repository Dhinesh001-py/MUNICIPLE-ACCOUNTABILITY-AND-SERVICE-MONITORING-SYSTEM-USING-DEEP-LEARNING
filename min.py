from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify
import pandas as pd
from datetime import datetime
from pathlib import Path
import os
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from openpyxl.styles import Font
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.sequence import pad_sequences
import pickle
import numpy as np

app = Flask(__name__)
app.secret_key = "secret-key"

# ====================== FILE PATHS ======================
DATAFILE = Path("data/complaints.xlsx")
UPLOAD_FOLDER = Path("static/uploads")
UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

MODEL_PATH = Path("backend/ml/priority_model.h5")
TOKENIZER_PATH = Path("backend/ml/tokenizer.pkl")
ENCODER_PATH = Path("backend/ml/label_encoder.pkl")

EXCEL_COLUMNS = [
    "id", "name", "address", "category", "description",
    "image_url", "status", "priority", "created_at"
]

# ====================== INITIAL SETUP ======================
def init_datafile():
    if not DATAFILE.exists():
        DATAFILE.parent.mkdir(parents=True, exist_ok=True)
        df = pd.DataFrame(columns=EXCEL_COLUMNS)
        df.to_excel(DATAFILE, index=False, engine="openpyxl")
        print("✅ Excel file created successfully!")

init_datafile()

# Load ML model (if available)
def load_priority_model():
    if MODEL_PATH.exists() and TOKENIZER_PATH.exists() and ENCODER_PATH.exists():
        try:
            model = load_model(MODEL_PATH)
            with open(TOKENIZER_PATH, "rb") as f:
                tokenizer = pickle.load(f)
            with open(ENCODER_PATH, "rb") as f:
                label_encoder = pickle.load(f)
            print("✅ Priority model loaded successfully!")
            return model, tokenizer, label_encoder
        except Exception as e:
            print("⚠️ Failed to load model/tokenizer/encoder:", e)
            return None, None, None
    else:
        print("⚠️ Priority model not found — using default priority.")
        return None, None, None

priority_model, tokenizer, label_encoder = load_priority_model()

# ====================== HELPERS ======================
def read_dataframe():
    """Read excel file into dataframe and normalize id column to int if possible."""
    df = pd.read_excel(DATAFILE, engine="openpyxl")
    if "id" in df.columns:
        # convert to integer ids consistently (if values present)
        try:
            # If there are missing ids, fill with 0 temporarily
            df["id"] = pd.to_numeric(df["id"], errors="coerce").fillna(0).astype(int)
        except Exception:
            # fallback: convert row index to id (shouldn't happen normally)
            df["id"] = df.index.astype(int)
    else:
        df["id"] = df.index.astype(int)
    # Ensure priority and status columns exist
    if "priority" not in df.columns:
        df["priority"] = "Medium"
    if "status" not in df.columns:
        df["status"] = "New"
    return df

def write_dataframe(df):
    """Write DataFrame to excel and keep hyperlink style for image_url column."""
    df.to_excel(DATAFILE, index=False, engine="openpyxl")
    # Re-open and make image_url column clickable (col F = 6th column)
    try:
        wb = load_workbook(DATAFILE)
        ws = wb.active
        # start from row 2 (assuming header row)
        for row_idx, val in enumerate(df["image_url"].fillna(""), start=2):
            cell = ws[f"F{row_idx}"]  # F column
            if val:
                cell.hyperlink = val
                cell.font = Font(color="0000FF", underline="single")
        wb.save(DATAFILE)
    except Exception as e:
        print("Warning: couldn't write hyperlinks:", e)

# ====================== APPEND TO EXCEL ======================
def append_complaint(record):
    df = read_dataframe()
    df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
    write_dataframe(df)

# ====================== ROUTES ======================
@app.route("/")
def index():
    categories = [
        "Garbage Dumping", "Streetlight Problem", "Women Safety",
        "Water Supply", "Road Damage", "Others"
    ]
    return render_template("index.html", categories=categories)

@app.route("/submit", methods=["POST"])
def submit():
    name = request.form.get("name")
    address = request.form.get("address")
    category = request.form.get("category")
    description = request.form.get("description")
    image_file = request.files.get("image_file")
    image_url = request.form.get("image_url", "").strip()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if not name or not address or not description:
        flash("Please fill in all required fields.", "danger")
        return redirect(url_for("index"))

    # Save uploaded image locally if provided
    if image_file and image_file.filename:
        filename = f"{int(datetime.now().timestamp())}_{secure_filename(image_file.filename)}"
        file_path = UPLOAD_FOLDER / filename
        image_file.save(file_path)
        image_url = url_for("static", filename=f"uploads/{filename}", _external=True)

    # Predict priority using model
    priority = "Medium"  # Default
    try:
        if priority_model and tokenizer is not None and label_encoder is not None:
            seq = tokenizer.texts_to_sequences([description])
            padded = pad_sequences(seq, maxlen=50)
            pred = priority_model.predict(padded)
            priority = label_encoder.inverse_transform([np.argmax(pred)])[0]
    except Exception as e:
        print("Priority prediction failed:", e)
        priority = "Medium"

    record = {
        "id": int(datetime.now().timestamp() * 1000),
        "name": name,
        "address": address,
        "category": category,
        "description": description,
        "image_url": image_url,
        "status": "New",
        "priority": priority,
        "created_at": created_at
    }

    append_complaint(record)
    flash(f"Complaint submitted successfully! (Priority: {priority})", "success")
    return redirect(url_for("dashboard"))

@app.route("/dashboard")
def dashboard():
    df = read_dataframe()

    # Assign priority based on category (auto rules) only if missing or default
    priority_map = {
        "Women Safety": "High",
        "Water Supply": "High",
        "Road Damage": "Medium",
        "Streetlight Problem": "Medium",
        "Garbage Dumping": "Low",
        "Others": "Low"
    }

    # Only override if category rule is stronger or priority is missing
    def compute_priority(row):
        cat_pr = priority_map.get(row.get("category"), None)
        if cat_pr:
            return cat_pr
        return row.get("priority", "Medium")

    if not df.empty:
        df["priority"] = df.apply(lambda x: compute_priority(x), axis=1)

    # Sort by priority (High → Medium → Low)
    priority_order = {"High": 1, "Medium": 2, "Low": 3}
    df["priority_order"] = df["priority"].map(priority_order).fillna(99)
    df = df.sort_values(by=["priority_order", "created_at"], ascending=[True, False])

    # Prepare rows for rendering (convert NaN to empty strings)
    rows = df.fillna("").to_dict(orient="records")

    # Compute counts for graphs
    priority_counts = df["priority"].value_counts().to_dict()
    category_counts = df["category"].value_counts().to_dict()
    status_counts = df["status"].value_counts().to_dict()

    return render_template(
        "dashboard.html",
        rows=rows,
        priority_counts=priority_counts,
        category_counts=category_counts,
        status_counts=status_counts
    )

# Classic form-based update (redirects back)
@app.route("/update_status/<int:cid>", methods=["POST"])
def update_status(cid):
    new_status = request.form.get("status")
    if not new_status:
        flash("No status provided", "warning")
        return redirect(url_for("dashboard"))

    df = read_dataframe()
    mask = df["id"] == int(cid)
    if mask.any():
        df.loc[mask, "status"] = new_status
        write_dataframe(df)
        flash("Status updated", "success")
    else:
        flash("Complaint not found", "danger")
    return redirect(url_for("dashboard"))

# JSON/AJAX endpoint for updating status (returns counts and updated row info)
@app.route("/api/update_status", methods=["POST"])
def api_update_status():
    data = request.get_json(silent=True) or request.form
    try:
        cid = int(data.get("id"))
        new_status = data.get("status")
    except Exception:
        return jsonify({"success": False, "error": "Invalid payload"}), 400

    if new_status not in ("New", "In Progress", "Resolved"):
        return jsonify({"success": False, "error": "Invalid status"}), 400

    df = read_dataframe()
    mask = df["id"] == cid
    if not mask.any():
        return jsonify({"success": False, "error": "Complaint not found"}), 404

    df.loc[mask, "status"] = new_status
    write_dataframe(df)

    # recompute counts
    status_counts = df["status"].value_counts().to_dict()
    # return updated row data to the client
    updated_row = df[mask].iloc[0].to_dict()
    return jsonify({
        "success": True,
        "row": updated_row,
        "status_counts": status_counts
    }), 200

# ====================== DOWNLOAD ======================
@app.route("/download")
def download():
    if not DATAFILE.exists():
        df = pd.DataFrame(columns=EXCEL_COLUMNS)
        df.to_excel(DATAFILE, index=False, engine="openpyxl")
    return send_file(
        str(DATAFILE),
        as_attachment=True,
        download_name="complaints.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ====================== RUN ======================
if __name__ == "__main__":
    app.run(debug=True)
